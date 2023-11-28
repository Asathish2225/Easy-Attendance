import json
import sys
import openpyxl  # To use Excel
from jsonrpclib.SimpleJSONRPCServer import SimpleJSONRPCServer
from openpyxl import Workbook
from datetime import datetime, time
import pytz  # Import the pytz library for working with time zones
import Server_UI
import os


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

entered_username = ""
entered_password = ""

# Location path for the Excel to be stored
folder_path = resource_path("C:\\Attendance\\")

# Define the IST time zone
ist = pytz.timezone('Asia/Kolkata')

# Define the start and end times for each period
period_timings = [
    {"name": "Hour 1", "start_time": time(8, 30), "end_time": time(9, 20)},
    {"name": "Hour 2", "start_time": time(9, 20), "end_time": time(10, 10)},
    {"name": "Hour 3", "start_time": time(10, 10), "end_time": time(11, 0)},
    {"name": "Hour 4", "start_time": time(11, 0), "end_time": time(11, 50)},
    {"name": "Hour 5", "start_time": time(11, 50), "end_time": time(12, 40)},
    {"name": "Hour 6", "start_time": time(12, 40), "end_time": time(13, 30)},
    {"name": "Hour 7", "start_time": time(13, 30), "end_time": time(14, 15)},
    {"name": "Hour 8", "start_time": time(14, 15), "end_time": time(15, 0)},
    {"name": "Hour 9", "start_time": time(15, 0), "end_time": time(15, 45)},
]


# Function to get the current period based on current time
def get_current_period():
    current_time = datetime.now().time()
    for period in period_timings:
        if period["start_time"] <= current_time <= period["end_time"]:

            return period["name"]
    return "No Period"  # If no period is currently active


# Function to generate a filename for the current period
def get_period_filename(period):
    # current_time = datetime.now()
    filename = f"{period}.xlsx"

    # Combine the folder path and filename to get the full file path
    full_path = folder_path + filename

    return full_path


# Create a new Excel workbook and add a worksheet for the current period
def create_day_workbook(date):
    workbook = Workbook()
    for period in period_timings:
        hour = period["name"]
        worksheet = workbook.create_sheet(title=hour)
        worksheet.append(["Room.No", "Name", "Reg.No", "Department", "Semester", "Section", "Subject", "Date", "Time","System Name"])  # Add headers to each sheet
        worksheet.title = hour  # Hour as sheet name
    del workbook["Sheet"]  # Remove the default sheet
    return workbook


# Function to add attendance for the current period
def add_attendance(room_number, name, reg_number, department,selected_semester, section, subject, client_hostname):
    current_period = get_current_period()

    if current_period == "No Period":
        return "No active period for attendance."

    current_time = datetime.now()
    period_filename = get_period_filename(current_time.strftime('%d-%m-%Y'))

    # If the workbook for the current day doesn't exist, create it
    if not os.path.exists(period_filename):
        workbook = create_day_workbook(current_time.strftime('%d-%m-%Y'))
    else:
        # Open the existing workbook
        workbook = openpyxl.load_workbook(period_filename)

    # Check if reg_number already exists in the current period's worksheet
    worksheet = workbook[current_period]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[2] == reg_number:  # Assuming reg_number is in the second column
            return f"Attendance for {reg_number} already recorded for {current_period}"

    # Check if the system name already exists in the worksheet
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=10, max_col=10):
        for cell in row:
            # print(f"Cell value: {cell.value}, System Name: {client_hostname}")
            if cell.value == client_hostname:
                return "Attendance for this system has already been recorded in this period."

    # Add data to the Excel worksheet
    worksheet.append([room_number, name, reg_number, department, selected_semester, section, subject, current_time.strftime('%d-%m-%Y'), current_time.strftime('%H:%M:%S'), client_hostname])

    # Save the workbook for the current day
    workbook.save(period_filename)

    # This will display the details in the server console
    print(f"Attendance recorded for {name} ({reg_number}) at {current_time.strftime('%H:%M:%S')} for {current_period}")

    return "Attendance has been successfully submitted."

script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the full path to the admin_credentials.json file
# This file contains the Username and Password

config_file_path = resource_path('admin_credentials.json')


# Function to check if the admin account exists in the credentials file
def check_admin_exists():
    try:
        with open(config_file_path, 'r') as config_file:
            config_data = json.load(config_file)
            return 'admin' in config_data
    except FileNotFoundError:
        return False


# Function to save admin credentials to the credentials file
def save_admin_credentials(credentials):
    try:
        with open(config_file_path, 'r') as config_file:
            config_data = json.load(config_file)
    except FileNotFoundError:
        config_data = {}

    config_data['admin'] = {
        'username': credentials['username'],
        'password': credentials['password']
    }

    with open(config_file_path, 'w') as file:
        json.dump(config_data, file)


# Function to check admin credentials
def check_admin_credentials(username, password):
    with open(config_file_path, 'r') as file:
        config_data = json.load(file)
    admin_data = config_data.get('admin', {})
    return admin_data.get('username') == username and admin_data.get('password') == password


# Define a callback function to handle credentials
def handle_credentials(username, password):
    global entered_username, entered_password  # Declare them as global variables
    entered_username = username
    entered_password = password


def main():
    global entered_username, entered_password

    # Call the UI function to obtain admin credentials
    Server_UI.get_admin_credentials(handle_credentials)

    # Admin account already exists, verify the credentials
    if check_admin_credentials(entered_username, entered_password):
        print("Admin login successful.")

        # Create the server
        server = SimpleJSONRPCServer(("10.1.121.93", 2509))
        # Register the add_attendance function
        server.register_function(add_attendance, 'add_attendance')
        print("Server is ready to accept connections.")
        server.serve_forever()
    else:
        # Do not start the server if admin login is unsuccessful
        print("Invalid admin credentials. Server will not start.")
        pass


if __name__ == "__main__":
    main()
